from __future__ import annotations

import csv
import io
import re
from pathlib import Path

import openpyxl
import requests

OUT = Path('data/subtel_sector_series/core_sheet_preview.csv')

SOURCES = [
    ('fixed_connections', 'https://www.subtel.gob.cl/wp-content/uploads/2026/05/1_SERIES_CONEXIONES_INTERNET_FIJA_MAR26_040526.xlsx', ['7.1.CO_TOT_FIJAS','7.7.CO_TEC_FIJAS']),
    ('mobile_connections', 'https://www.subtel.gob.cl/wp-content/uploads/2026/05/2_SERIES_CONEXIONES_INTERNET_MO%CC%81VIL-MAR26-040526.xlsx', ['8.1.CO_TEC_MOVIL']),
    ('mobile_traffic', 'https://www.subtel.gob.cl/wp-content/uploads/2026/05/3_SERIES_TRAFICO_DATOS_MOVILES-MAR26-040526.xlsx', ['9.1.TRAF_SENT']),
    ('fixed_traffic', 'https://www.subtel.gob.cl/wp-content/uploads/2026/05/3_SERIES_TRAFICO_DATOS_FIJOS-MAR26-040526.xlsx', ['10.1.TRAF_SENT']),
]


def clean(v):
    if v is None: return ''
    if hasattr(v, 'isoformat') and not isinstance(v, str):
        try: return v.isoformat()
        except Exception: pass
    return re.sub(r'\s+', ' ', str(v)).strip()


def download(url):
    r=requests.get(url,timeout=180); r.raise_for_status()
    if not r.content.startswith(b'PK'): raise RuntimeError(f'Not xlsx: {url}')
    return r.content


def main():
    OUT.parent.mkdir(parents=True,exist_ok=True)
    rows=[]
    for dataset,url,sheets in SOURCES:
        wb=openpyxl.load_workbook(io.BytesIO(download(url)),read_only=True,data_only=True)
        for sheet in sheets:
            ws=wb[sheet]
            kept=0
            for row_no,row in enumerate(ws.iter_rows(min_row=1,max_row=min(ws.max_row,40),max_col=min(ws.max_column,25),values_only=True),start=1):
                vals=[clean(v) for v in row]
                if not any(vals): continue
                rows.append({'dataset':dataset,'sheet':sheet,'row':row_no,'values':' || '.join(vals)})
                kept+=1
                if kept>=25: break
    with OUT.open('w',encoding='utf-8',newline='') as fh:
        w=csv.DictWriter(fh,fieldnames=['dataset','sheet','row','values']); w.writeheader(); w.writerows(rows)
    print('preview_rows',len(rows))

if __name__=='__main__': main()
