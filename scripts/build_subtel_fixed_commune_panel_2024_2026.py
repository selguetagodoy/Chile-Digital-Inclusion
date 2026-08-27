#!/usr/bin/env python3
"""Build a validated June-2024 / June-2025 / June-2026 fixed-Internet commune panel.

Uses the latest June-2026 SUBTEL workbook as the vintage for all three historical
cuts, applies the same regional formula-block reconciliation used by the validated
June-2026 commune extractor, and joins Censo 2024 household/rurality denominators.
Historical values therefore use one consistent latest workbook vintage.
"""
from __future__ import annotations

import csv
import io
import urllib.request
from pathlib import Path

import openpyxl

import build_subtel_fixed_commune_2026 as base

URL = 'https://www.subtel.gob.cl/wp-content/uploads/2026/08/1_SERIES_CONEXIONES_INTERNET_FIJA_JUN26_100826.xlsx'
PERIODS = ['2024-06', '2025-06', '2026-06']
SHEETS = {
    'total': '7.11.CO_FIJAS_COMUNA',
    'residential': '7.11.1.CO_FIJAS_RES_COMUNA',
}
CENSO = Path('data/censo_2024/communes_connectivity_2024.csv')
NATIONAL = Path('data/subtel_sector_series/fixed_connections_monthly.csv')
OUT = Path('data/fixed_infrastructure_2026')

MONTH_ALIASES = {
    1: {'ene', 'enero'}, 2: {'feb', 'febrero'}, 3: {'mar', 'marzo'},
    4: {'abr', 'abril'}, 5: {'may', 'mayo'}, 6: {'jun', 'junio'},
    7: {'jul', 'julio'}, 8: {'ago', 'agosto'}, 9: {'sep', 'sept', 'septiembre'},
    10: {'oct', 'octubre'}, 11: {'nov', 'noviembre'}, 12: {'dic', 'diciembre'},
}


def locate_period_column(ws, target_period: str):
    year, month = map(int, target_period.split('-'))
    current_year = None
    matches = []
    for col in range(1, ws.max_column + 1):
        raw_year = ws.cell(8, col).value
        if isinstance(raw_year, (int, float)) and int(raw_year) == raw_year and 2000 <= int(raw_year) <= 2100:
            current_year = int(raw_year)
        month_label = base.clean(ws.cell(9, col).value).lower().replace('.', '')
        if current_year == year and month_label in MONTH_ALIASES[month]:
            matches.append(col)
    if not matches:
        raise RuntimeError(f'Could not locate {target_period} in {ws.title}')
    return max(matches)


def pct_change(a, b):
    if a in (None, 0) or b is None:
        return None
    return (b / a - 1) * 100


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    request = urllib.request.Request(URL, headers={'User-Agent': 'Mozilla/5.0 Chile-Digital-Inclusion/1.0'})
    with urllib.request.urlopen(request, timeout=180) as response:
        content = response.read()
    wb_values = openpyxl.load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    wb_formulas = openpyxl.load_workbook(io.BytesIO(content), read_only=False, data_only=False)

    with base.COMMUNES.open(encoding='utf-8') as fh:
        catalogue = list(csv.DictReader(fh))
    catalogue_by_region = {}
    for region in base.REGION_CODES:
        region_rows = [r for r in catalogue if int(r['region']) == region]
        catalogue_by_region[region] = sorted(region_rows, key=lambda r: base.source_sort_key(r['comuna_nombre']))

    with CENSO.open(encoding='utf-8') as fh:
        censo = {int(r['comuna']): r for r in csv.DictReader(fh)}

    with NATIONAL.open(encoding='utf-8') as fh:
        national = {r['period']: int(round(float(r['fixed_connections']))) for r in csv.DictReader(fh) if r['period'] in PERIODS}
    if set(national) != set(PERIODS):
        raise RuntimeError(f'Missing national controls for {sorted(set(PERIODS) - set(national))}')

    maps = {}
    alignment_all = []
    provenance_all = []
    manifest = []

    for target_period in PERIODS:
        for metric, sheet_name in SHEETS.items():
            ws_values = wb_values[sheet_name]
            ws_formulas = wb_formulas[sheet_name]
            col = locate_period_column(ws_values, target_period)
            key = f'{metric}_{target_period}'
            mapped, alignment, provenance, issues = base.build_metric(
                ws_values, ws_formulas, col, catalogue_by_region, key
            )
            if issues:
                raise RuntimeError(f'Reconciliation failed for {key}: {issues}')
            maps[key] = mapped
            for row in alignment:
                row['period'] = target_period
                alignment_all.append(row)
            for row in provenance:
                row['period'] = target_period
                provenance_all.append(row)
            manifest.append({
                'period': target_period, 'metric': metric, 'source_sheet': sheet_name,
                'source_column_number': col,
                'mapping_method': 'regional_formula_block_order_with_subtotal_reconciliation',
                'source_url': URL,
                'vintage': '2026-06 workbook',
            })

        numeric_total = sum(maps[f'total_{target_period}'].values())
        if numeric_total != national[target_period]:
            raise RuntimeError(f'{target_period}: commune sum {numeric_total} != national {national[target_period]}')

    rows = []
    by_code = {int(r['comuna']): r for r in catalogue}
    for code in sorted(by_code):
        geo = by_code[code]
        census = censo[code]
        households = int(float(census['hogares_total']))
        rural_pct = float(census['hogares_rurales_pct'])
        row = {
            'region': geo['region'], 'region_nombre': geo['region_nombre'],
            'provincia': geo['provincia'], 'provincia_nombre': geo['provincia_nombre'],
            'comuna': geo['comuna'], 'comuna_nombre': geo['comuna_nombre'],
            'censo_2024_households': households,
            'censo_2024_rural_households_pct': round(rural_pct, 4),
        }
        for target_period in PERIODS:
            suffix = target_period.replace('-', 'm')
            total = maps[f'total_{target_period}'].get(code)
            residential = maps[f'residential_{target_period}'].get(code)
            status = 'reported' if total is not None and residential is not None else 'source_blank'
            row[f'fixed_total_{suffix}'] = '' if total is None else total
            row[f'fixed_residential_{suffix}'] = '' if residential is None else residential
            row[f'fixed_residential_per_100_censo_households_{suffix}'] = '' if residential is None or not households else round(residential / households * 100, 4)
            row[f'source_status_{suffix}'] = status

        for base_period, end_period, label in [
            ('2024-06', '2025-06', '2024m06_to_2025m06'),
            ('2025-06', '2026-06', '2025m06_to_2026m06'),
            ('2024-06', '2026-06', '2024m06_to_2026m06'),
        ]:
            a = maps[f'residential_{base_period}'].get(code)
            b = maps[f'residential_{end_period}'].get(code)
            row[f'fixed_residential_change_{label}'] = '' if a is None or b is None else b - a
            pc = pct_change(a, b)
            row[f'fixed_residential_growth_pct_{label}'] = '' if pc is None else round(pc, 4)
            ia = a / households * 100 if a is not None and households else None
            ib = b / households * 100 if b is not None and households else None
            row[f'fixed_residential_intensity_change_pp_{label}'] = '' if ia is None or ib is None else round(ib - ia, 4)
        rows.append(row)

    panel_path = OUT / 'commune_fixed_connectivity_panel_2024_2026.csv'
    with panel_path.open('w', newline='', encoding='utf-8') as fh:
        w = csv.DictWriter(fh, fieldnames=list(rows[0].keys()))
        w.writeheader(); w.writerows(rows)

    # Scatter/ranking-ready subset: suppress very small household bases from rank tables,
    # while retaining all 346 communes in the canonical panel.
    ranked = [r for r in rows if r['source_status_2026m06'] == 'reported' and r['censo_2024_households'] >= 1000]
    ranked.sort(key=lambda r: float(r['fixed_residential_per_100_censo_households_2026m06']))
    gap_rows = []
    for rank, r in enumerate(ranked[:30], start=1):
        gap_rows.append({
            'rank_lowest_intensity_2026': rank,
            'comuna': r['comuna'], 'comuna_nombre': r['comuna_nombre'], 'region_nombre': r['region_nombre'],
            'censo_2024_households': r['censo_2024_households'],
            'censo_2024_rural_households_pct': r['censo_2024_rural_households_pct'],
            'fixed_residential_per_100_censo_households_2024m06': r['fixed_residential_per_100_censo_households_2024m06'],
            'fixed_residential_per_100_censo_households_2025m06': r['fixed_residential_per_100_censo_households_2025m06'],
            'fixed_residential_per_100_censo_households_2026m06': r['fixed_residential_per_100_censo_households_2026m06'],
            'fixed_residential_growth_pct_2025m06_to_2026m06': r['fixed_residential_growth_pct_2025m06_to_2026m06'],
            'ranking_scope': 'communes with >=1000 Censo 2024 households and reported June-2026 SUBTEL value',
        })
    with (OUT / 'commune_fixed_low_intensity_rank_2026.csv').open('w', newline='', encoding='utf-8') as fh:
        w = csv.DictWriter(fh, fieldnames=list(gap_rows[0].keys()))
        w.writeheader(); w.writerows(gap_rows)

    align_fields = ['period'] + [k for k in alignment_all[0].keys() if k != 'period']
    base.write_dict_csv(OUT / 'source_alignment_qa_panel_2024_2026.csv', align_fields, alignment_all)
    prov_fields = ['period'] + [k for k in provenance_all[0].keys() if k != 'period']
    base.write_dict_csv(OUT / 'source_row_mapping_panel_2024_2026.csv', prov_fields, provenance_all)
    with (OUT / 'extraction_manifest_panel_2024_2026.csv').open('w', newline='', encoding='utf-8') as fh:
        w = csv.DictWriter(fh, fieldnames=list(manifest[0].keys()))
        w.writeheader(); w.writerows(manifest)

    print('panel_rows', len(rows))
    print('rank_rows', len(gap_rows))
    print('alignment_checks', len(alignment_all))
    for p in PERIODS:
        print(p, 'national_control', national[p], 'commune_sum', sum(maps[f'total_{p}'].values()))


if __name__ == '__main__':
    main()
