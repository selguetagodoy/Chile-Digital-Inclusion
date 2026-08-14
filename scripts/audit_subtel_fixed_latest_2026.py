from __future__ import annotations

import csv
import io
import re
from pathlib import Path

import openpyxl
import requests

URL='https://www.subtel.gob.cl/wp-content/uploads/2026/05/1_SERIES_CONEXIONES_INTERNET_FIJA_MAR26_040526.xlsx'
OUT=Path('data/subtel_sector_series/latest_fixed_technology_audit.csv')
REGIONAL_OUT=Path('data/subtel_sector_series/latest_fixed_technology_regional_audit.csv')


def val(v):
    if v is None:
        return ''
    return re.sub(r'\s+', ' ', str(v)).strip()


def integer(v):
    if v in (None, '', '---'):
        return None
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return None


def main():
    r=requests.get(URL,timeout=180)
    r.raise_for_status()
    wb=openpyxl.load_workbook(io.BytesIO(r.content),read_only=True,data_only=True)

    # Keep a compact audit of the legacy national sheet. Its recent taxonomy is
    # not treated as a homogeneous longitudinal technology series.
    ws=wb['7.7.CO_TEC_FIJAS']
    rows=[]
    for row_no,row in enumerate(ws.iter_rows(min_row=1,max_row=ws.max_row,max_col=10,values_only=True),start=1):
        vals=list(row)
        if row_no<=8 or any('2026' in val(v) for v in vals[:3]):
            rows.append({'source_row':row_no,**{f'col{i+1}':val(vals[i]) for i in range(10)}})
    tail=[]
    for row_no,row in enumerate(
        ws.iter_rows(min_row=max(1,ws.max_row-12),max_row=ws.max_row,max_col=10,values_only=True),
        start=max(1,ws.max_row-12),
    ):
        vals=list(row)
        if any(val(v) for v in vals):
            tail.append({'source_row':row_no,**{f'col{i+1}':val(vals[i]) for i in range(10)}})
    seen={r['source_row'] for r in rows}
    rows.extend(r for r in tail if r['source_row'] not in seen)
    OUT.parent.mkdir(parents=True,exist_ok=True)
    fields=['source_row']+[f'col{i}' for i in range(1,11)]
    with OUT.open('w',encoding='utf-8',newline='') as fh:
        w=csv.DictWriter(fh,fieldnames=fields)
        w.writeheader(); w.writerows(rows)

    # March 2026 regional technology is published in 7.7.1. The previous audit
    # referenced a non-existent 7.6 sheet. Read only explicit labelled totals,
    # the same defensible contract used by the canonical sector builder.
    ws=wb['7.7.1.CO_TEC_RG_EMP_FIJAS']
    headers=[val(v) for v in next(ws.iter_rows(min_row=8,max_row=8,max_col=ws.max_column,values_only=True))]
    by_label={}
    for i,label in enumerate(headers):
        by_label.setdefault(label,[]).append(i)

    def one(label):
        hits=by_label.get(label,[])
        if len(hits)!=1:
            raise RuntimeError(f'Expected one column {label!r}, found {hits}')
        return hits[0]

    region_i=one('Región')
    tech_cols={
        'adsl_connections': one('Total Conexiones ADSL'),
        'hfc_connections': one('Total Conexiones HFC (Cable Modem)'),
        'wimax_connections': one('Total Conexiones WIMAX'),
        'fttx_fiber_connections': one('Total Conexiones FTTX'),
    }
    grand_hits=by_label.get('Total Conexiones',[])
    if not grand_hits:
        raise RuntimeError('No explicit Total Conexiones column in 7.7.1')
    grand_i=grand_hits[-1]

    regional=[]
    national_total=None
    national_row=None
    for row_no,row in enumerate(ws.iter_rows(min_row=9,max_col=ws.max_column,values_only=True),start=9):
        values=list(row)
        region=val(values[region_i]) if region_i < len(values) else ''
        grand=integer(values[grand_i]) if grand_i < len(values) else None
        if region.lower()=='total':
            national_total=grand
            national_row=row_no
            continue
        try:
            region_no=int(float(region))
        except (TypeError,ValueError):
            continue
        if not 1<=region_no<=16 or grand is None:
            continue
        counts={name:integer(values[idx]) or 0 for name,idx in tech_cols.items()}
        known=sum(counts.values())
        residual=grand-known
        if residual<0:
            raise RuntimeError(f'Region {region_no}: explicit technology totals exceed grand total')
        regional.append({
            'period':'2026-03',
            'region':region_no,
            **counts,
            'other_fixed_technologies_residual':residual,
            'total_fixed_connections':grand,
            'source_sheet':'7.7.1.CO_TEC_RG_EMP_FIJAS',
            'source_row':row_no,
        })

    if len(regional)!=16:
        raise RuntimeError(f'Expected 16 regional rows, found {len(regional)}')
    regional_sum=sum(r['total_fixed_connections'] for r in regional)
    if national_total is None:
        raise RuntimeError('Could not locate national Total row in 7.7.1')
    if regional_sum!=national_total:
        raise RuntimeError(f'Regional totals do not tie to national total: {regional_sum} != {national_total}')

    regional.sort(key=lambda r:r['region'])
    regional_fields=[
        'period','region','adsl_connections','hfc_connections','wimax_connections',
        'fttx_fiber_connections','other_fixed_technologies_residual',
        'total_fixed_connections','source_sheet','source_row',
    ]
    with REGIONAL_OUT.open('w',encoding='utf-8',newline='') as fh:
        w=csv.DictWriter(fh,fieldnames=regional_fields)
        w.writeheader(); w.writerows(regional)

    print('legacy_national_audit_rows',len(rows))
    print('current_regional_rows',len(regional))
    print('current_regional_total',regional_sum)
    print('current_national_total',national_total,'source_row',national_row)


if __name__=='__main__':
    main()
