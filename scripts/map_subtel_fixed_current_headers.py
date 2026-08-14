from __future__ import annotations

import csv
import io
from pathlib import Path

import openpyxl
import requests
from openpyxl.utils import get_column_letter

URL='https://www.subtel.gob.cl/wp-content/uploads/2026/05/1_SERIES_CONEXIONES_INTERNET_FIJA_MAR26_040526.xlsx'
OUT=Path('data/subtel_sector_series/current_fixed_technology_header_map.csv')
TOTALS=Path('data/subtel_sector_series/current_fixed_technology_march2026_totals.csv')
SHEET='7.7.1.CO_TEC_RG_EMP_FIJAS'


def clean(v):
    return '' if v is None else str(v).strip().replace('\n',' ')


def main():
    r=requests.get(URL,timeout=180); r.raise_for_status()
    wb=openpyxl.load_workbook(io.BytesIO(r.content),read_only=True,data_only=True)
    ws=wb[SHEET]
    group=list(next(ws.iter_rows(min_row=7,max_row=7,max_col=ws.max_column,values_only=True)))
    sub=list(next(ws.iter_rows(min_row=8,max_row=8,max_col=ws.max_column,values_only=True)))
    total=list(next(ws.iter_rows(min_row=25,max_row=25,max_col=ws.max_column,values_only=True)))

    # Fill merged-style group labels horizontally only for audit/display.
    current=''
    rows=[]
    total_rows=[]
    for i in range(ws.max_column):
        if clean(group[i]): current=clean(group[i])
        row={
            'column':i+1,
            'excel_column':get_column_letter(i+1),
            'group_header':current,
            'column_header':clean(sub[i]),
            'march_2026_total_row_value':clean(total[i]),
        }
        rows.append(row)
        h=clean(sub[i]).lower()
        if h.startswith('total conexiones') or h in {'total de conexiones fijas','total conexiones fijas'}:
            total_rows.append(row)

    OUT.parent.mkdir(parents=True,exist_ok=True)
    fields=['column','excel_column','group_header','column_header','march_2026_total_row_value']
    with OUT.open('w',encoding='utf-8',newline='') as fh:
        w=csv.DictWriter(fh,fieldnames=fields); w.writeheader(); w.writerows(rows)
    with TOTALS.open('w',encoding='utf-8',newline='') as fh:
        w=csv.DictWriter(fh,fieldnames=fields); w.writeheader(); w.writerows(total_rows)

    print('columns',ws.max_column,'total_columns',len(total_rows))
    for row in total_rows: print(row)

if __name__=='__main__': main()
