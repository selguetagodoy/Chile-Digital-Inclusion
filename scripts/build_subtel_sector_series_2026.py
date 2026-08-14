from __future__ import annotations

import csv
import io
import math
from pathlib import Path

import openpyxl
import requests

OUTDIR = Path('data/subtel_sector_series')

SOURCES = {
    'fixed': 'https://www.subtel.gob.cl/wp-content/uploads/2026/05/1_SERIES_CONEXIONES_INTERNET_FIJA_MAR26_040526.xlsx',
    'mobile': 'https://www.subtel.gob.cl/wp-content/uploads/2026/05/2_SERIES_CONEXIONES_INTERNET_MO%CC%81VIL-MAR26-040526.xlsx',
    'mobile_traffic': 'https://www.subtel.gob.cl/wp-content/uploads/2026/05/3_SERIES_TRAFICO_DATOS_MOVILES-MAR26-040526.xlsx',
    'fixed_traffic': 'https://www.subtel.gob.cl/wp-content/uploads/2026/05/3_SERIES_TRAFICO_DATOS_FIJOS-MAR26-040526.xlsx',
}

MONTHS = {
    'ene': 1, 'feb': 2, 'mar': 3, 'abr': 4, 'may': 5, 'jun': 6,
    'jul': 7, 'ago': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dic': 12,
}


def download(url: str) -> bytes:
    r = requests.get(url, timeout=180, allow_redirects=True)
    r.raise_for_status()
    if not r.content.startswith(b'PK'):
        raise RuntimeError(f'Expected XLSX from {url}; type={r.headers.get("content-type")} bytes={len(r.content)}')
    return r.content


def workbook(key: str):
    return openpyxl.load_workbook(io.BytesIO(download(SOURCES[key])), read_only=True, data_only=True)


def number(v):
    if v is None or v == '' or v == '---':
        return None
    if isinstance(v, bool):
        return None
    try:
        x = float(v)
        return None if math.isnan(x) else x
    except (TypeError, ValueError):
        return None


def integer(v):
    x = number(v)
    return None if x is None else int(round(x))


def month_no(v):
    if v is None:
        return None
    return MONTHS.get(str(v).strip().lower().replace('.', '')[:3])


def period(year: int, month: int) -> str:
    return f'{year:04d}-{month:02d}'


def iter_period_rows(ws, start_row: int, year_col: int, month_col: int, max_col: int):
    current_year = None
    for row_no, row in enumerate(ws.iter_rows(min_row=start_row, max_col=max_col, values_only=True), start=start_row):
        vals = list(row)
        y = integer(vals[year_col - 1]) if year_col <= len(vals) else None
        if y is not None and 1990 <= y <= 2100:
            current_year = y
        m = month_no(vals[month_col - 1]) if month_col <= len(vals) else None
        if current_year is not None and m is not None:
            yield row_no, current_year, m, vals


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open('w', encoding='utf-8', newline='') as fh:
        w = csv.DictWriter(fh, fieldnames=fields)
        w.writeheader(); w.writerows(rows)


def build_fixed_total(wb):
    rows = []
    ws = wb['7.1.CO_TOT_FIJAS']
    for row_no, y, m, v in iter_period_rows(ws, 9, 3, 4, 8):
        total = integer(v[4])
        if total is None:
            continue
        rows.append({
            'period': period(y, m), 'year': y, 'month': m,
            'fixed_connections_total': total,
            'annual_growth_pct': None if number(v[5]) is None else round(number(v[5]) * 100, 6),
            'penetration_per_100_inhabitants': None if number(v[6]) is None else round(number(v[6]), 6),
            'penetration_annual_change_pp': None if number(v[7]) is None else round(number(v[7]), 6),
            'source_sheet': '7.1.CO_TOT_FIJAS', 'source_row': row_no,
        })
    return rows


def build_fixed_technology_snapshot(wb, expected_total: int):
    """Read current fixed technology only from explicit labelled totals.

    The legacy national history sheet 7.7 changes taxonomy/column meaning over
    time. It is deliberately not used as a homogeneous technology series.
    The current cross-sectional sheet 7.7.1 has explicit labelled total columns.
    """
    ws = wb['7.7.1.CO_TEC_RG_EMP_FIJAS']
    headers = list(next(ws.iter_rows(min_row=8, max_row=8, max_col=ws.max_column, values_only=True)))

    def clean(v):
        return '' if v is None else str(v).strip()

    header_to_indices: dict[str, list[int]] = {}
    for i, h in enumerate(headers):
        header_to_indices.setdefault(clean(h), []).append(i)

    required = {
        'ADSL': 'Total Conexiones ADSL',
        'HFC': 'Total Conexiones HFC (Cable Modem)',
        'WIMAX': 'Total Conexiones WIMAX',
        'FTTX_FIBER': 'Total Conexiones FTTX',
    }
    indices = {}
    for key, label in required.items():
        candidates = header_to_indices.get(label, [])
        if len(candidates) != 1:
            raise RuntimeError(f'Expected one current fixed-tech column {label}, found {candidates}')
        indices[key] = candidates[0]

    grand_candidates = header_to_indices.get('Total Conexiones', [])
    if not grand_candidates:
        raise RuntimeError('No explicit Total Conexiones column in 7.7.1')
    grand_idx = grand_candidates[-1]

    total_row = None
    total_row_no = None
    for row_no, row in enumerate(ws.iter_rows(min_row=9, max_col=ws.max_column, values_only=True), start=9):
        vals = list(row)
        if grand_idx < len(vals) and integer(vals[grand_idx]) == expected_total:
            total_row = vals
            total_row_no = row_no
            break
    if total_row is None:
        raise RuntimeError(f'Could not find 7.7.1 national total row matching {expected_total}')

    counts = {key: integer(total_row[idx]) for key, idx in indices.items()}
    if any(v is None for v in counts.values()):
        raise RuntimeError(f'Missing explicit fixed technology totals: {counts}')
    known = sum(counts.values())
    other = expected_total - known
    if other < 0:
        raise RuntimeError(f'Explicit technology totals exceed fixed total: known={known} total={expected_total}')

    rows = []
    for tech, value in [
        ('ADSL', counts['ADSL']),
        ('HFC', counts['HFC']),
        ('WIMAX', counts['WIMAX']),
        ('FTTX_FIBER', counts['FTTX_FIBER']),
        ('OTHER_FIXED_TECHNOLOGIES_RESIDUAL', other),
    ]:
        rows.append({
            'period': '2026-03',
            'technology': tech,
            'connections': value,
            'share_pct': round(value / expected_total * 100, 6) if expected_total else None,
            'source_sheet': '7.7.1.CO_TEC_RG_EMP_FIJAS',
            'source_row': total_row_no,
            'derivation': 'explicit labelled total' if tech != 'OTHER_FIXED_TECHNOLOGIES_RESIDUAL' else 'grand total minus ADSL/HFC/WIMAX/FTTX explicit totals',
        })
    rows.append({
        'period': '2026-03', 'technology': 'TOTAL_FIXED_CONNECTIONS',
        'connections': expected_total, 'share_pct': 100.0,
        'source_sheet': '7.7.1.CO_TEC_RG_EMP_FIJAS', 'source_row': total_row_no,
        'derivation': 'explicit Total Conexiones matching 7.1 national total',
    })
    return rows


def build_mobile(wb):
    rows = []
    ws = wb['8.1.CO_TEC_MOVIL']
    for row_no, y, m, v in iter_period_rows(ws, 8, 2, 3, 15):
        g2, g3, g4, g5 = [integer(v[i]) for i in range(3, 7)]
        total = integer(v[7])
        broadband = integer(v[8])
        if total is None:
            continue
        rows.append({
            'period': period(y, m), 'year': y, 'month': m,
            'connections_2g': g2, 'connections_3g': g3,
            'connections_4g': g4, 'connections_5g': g5,
            'mobile_connections_total': total,
            'connections_3g_4g_5g': broadband,
            'penetration_2g_per_100': None if number(v[9]) is None else round(number(v[9]), 6),
            'penetration_3g_per_100': None if number(v[10]) is None else round(number(v[10]), 6),
            'penetration_4g_per_100': None if number(v[11]) is None else round(number(v[11]), 6),
            'penetration_5g_per_100': None if number(v[12]) is None else round(number(v[12]), 6),
            'penetration_3g_4g_5g_per_100': None if number(v[13]) is None else round(number(v[13]), 6),
            'penetration_total_per_100': None if number(v[14]) is None else round(number(v[14]), 6),
            'source_sheet': '8.1.CO_TEC_MOVIL', 'source_row': row_no,
        })
    return rows


def build_traffic(wb, sheet: str, output_prefix: str):
    rows = []
    ws = wb[sheet]
    for row_no, y, m, v in iter_period_rows(ws, 11, 2, 3, 7):
        down, up, total, growth = number(v[3]), number(v[4]), number(v[5]), number(v[6])
        if total is None:
            continue
        rows.append({
            'period': period(y, m), 'year': y, 'month': m,
            f'{output_prefix}_downlink_tb': round(down, 6) if down is not None else None,
            f'{output_prefix}_uplink_tb': round(up, 6) if up is not None else None,
            f'{output_prefix}_total_tb': round(total, 6),
            'annual_growth_pct': None if growth is None else round(growth * 100, 6),
            'source_sheet': sheet, 'source_row': row_no,
        })
    return rows


def build_long_core(fixed_total, fixed_snapshot, mobile, mobile_traffic, fixed_traffic):
    rows = []

    def add(source_rows, metrics, series_group, units):
        for r in source_rows:
            for metric in metrics:
                value = r.get(metric)
                if value is None:
                    continue
                rows.append({
                    'period': r['period'],
                    'year': int(r.get('year') or r['period'][:4]),
                    'month': int(r.get('month') or r['period'][5:7]),
                    'series_group': series_group,
                    'indicator': metric,
                    'value': value,
                    'unit': units.get(metric, ''),
                    'source_sheet': r['source_sheet'],
                    'source_row': r['source_row'],
                })

    add(fixed_total, ['fixed_connections_total', 'penetration_per_100_inhabitants'], 'fixed_connections', {
        'fixed_connections_total': 'connections', 'penetration_per_100_inhabitants': 'per_100_inhabitants'})
    add(mobile, ['connections_2g', 'connections_3g', 'connections_4g', 'connections_5g', 'mobile_connections_total', 'connections_3g_4g_5g'], 'mobile_connections', {
        'connections_2g': 'connections', 'connections_3g': 'connections', 'connections_4g': 'connections',
        'connections_5g': 'connections', 'mobile_connections_total': 'connections', 'connections_3g_4g_5g': 'connections'})
    add(mobile_traffic, ['mobile_traffic_downlink_tb', 'mobile_traffic_uplink_tb', 'mobile_traffic_total_tb'], 'mobile_traffic', {
        'mobile_traffic_downlink_tb': 'TB', 'mobile_traffic_uplink_tb': 'TB', 'mobile_traffic_total_tb': 'TB'})
    add(fixed_traffic, ['fixed_traffic_downlink_tb', 'fixed_traffic_uplink_tb', 'fixed_traffic_total_tb'], 'fixed_traffic', {
        'fixed_traffic_downlink_tb': 'TB', 'fixed_traffic_uplink_tb': 'TB', 'fixed_traffic_total_tb': 'TB'})

    for r in fixed_snapshot:
        if r['technology'] == 'TOTAL_FIXED_CONNECTIONS':
            continue
        rows.append({
            'period': r['period'], 'year': 2026, 'month': 3,
            'series_group': 'fixed_technology_snapshot',
            'indicator': f"fixed_{r['technology'].lower()}_connections",
            'value': r['connections'], 'unit': 'connections',
            'source_sheet': r['source_sheet'], 'source_row': r['source_row'],
        })
    fiber = next(r for r in fixed_snapshot if r['technology'] == 'FTTX_FIBER')
    rows.append({
        'period': '2026-03', 'year': 2026, 'month': 3,
        'series_group': 'fixed_technology_snapshot', 'indicator': 'fixed_fttx_fiber_share_pct',
        'value': fiber['share_pct'], 'unit': 'percent',
        'source_sheet': fiber['source_sheet'], 'source_row': fiber['source_row'],
    })
    return rows


def build_qa(fixed_total, fixed_snapshot, mobile, mobile_traffic, fixed_traffic):
    groups = {
        'fixed_connections': fixed_total,
        'mobile_connections': mobile,
        'mobile_traffic': mobile_traffic,
        'fixed_traffic': fixed_traffic,
    }
    rows = []
    for name, data in groups.items():
        periods = [r['period'] for r in data]
        rows.extend([
            {'check': f'{name}_rows', 'value': len(data), 'expectation': 'positive'},
            {'check': f'{name}_unique_periods', 'value': len(set(periods)), 'expectation': str(len(data))},
            {'check': f'{name}_first_period', 'value': min(periods) if periods else '', 'expectation': 'source workbook effective range'},
            {'check': f'{name}_last_period', 'value': max(periods) if periods else '', 'expectation': '2026-03'},
        ])

    latest_fixed = next(r for r in fixed_total if r['period'] == '2026-03')
    jan_mobile = next(r for r in mobile if r['period'] == '2026-01')
    mar_mobile = next(r for r in mobile if r['period'] == '2026-03')
    fiber = next(r for r in fixed_snapshot if r['technology'] == 'FTTX_FIBER')
    total_snapshot = next(r for r in fixed_snapshot if r['technology'] == 'TOTAL_FIXED_CONNECTIONS')
    sector_snapshot_5g = 10_367_754

    rows.extend([
        {'check': '2026m03_fixed_total', 'value': latest_fixed['fixed_connections_total'], 'expectation': '4859679'},
        {'check': '2026m03_fixed_snapshot_total', 'value': total_snapshot['connections'], 'expectation': '4859679'},
        {'check': '2026m03_fixed_fttx_connections', 'value': fiber['connections'], 'expectation': '4147629 explicit labelled total'},
        {'check': '2026m03_fixed_fttx_share_pct', 'value': fiber['share_pct'], 'expectation': 'approximately 85.3'},
        {'check': '2026m01_5g_connections', 'value': jan_mobile['connections_5g'], 'expectation': '10161957 official January publication'},
        {'check': '2026m03_5g_connections_workbook', 'value': mar_mobile['connections_5g'], 'expectation': '10356448 official monthly XLSX'},
        {'check': '2026q1_5g_sector_snapshot_reference', 'value': sector_snapshot_5g, 'expectation': 'separate official sector snapshot retained elsewhere'},
        {'check': '2026m03_5g_difference_workbook_minus_snapshot', 'value': mar_mobile['connections_5g'] - sector_snapshot_5g, 'expectation': '-11306 provenance discrepancy; do not force reconciliation'},
    ])
    return rows


def main():
    OUTDIR.mkdir(parents=True, exist_ok=True)

    fixed_wb = workbook('fixed')
    fixed_total = build_fixed_total(fixed_wb)
    march_fixed_total = next(r for r in fixed_total if r['period'] == '2026-03')['fixed_connections_total']
    fixed_snapshot = build_fixed_technology_snapshot(fixed_wb, march_fixed_total)
    mobile = build_mobile(workbook('mobile'))
    mobile_traffic = build_traffic(workbook('mobile_traffic'), '9.1.TRAF_SENT', 'mobile_traffic')
    fixed_traffic = build_traffic(workbook('fixed_traffic'), '10.1.TRAF_SENT', 'fixed_traffic')

    write_csv(OUTDIR / 'fixed_connections_monthly.csv', fixed_total, list(fixed_total[0].keys()))
    write_csv(OUTDIR / 'fixed_technology_snapshot_2026_03.csv', fixed_snapshot, list(fixed_snapshot[0].keys()))
    write_csv(OUTDIR / 'mobile_connections_by_technology_monthly.csv', mobile, list(mobile[0].keys()))
    write_csv(OUTDIR / 'mobile_data_traffic_monthly.csv', mobile_traffic, list(mobile_traffic[0].keys()))
    write_csv(OUTDIR / 'fixed_data_traffic_monthly.csv', fixed_traffic, list(fixed_traffic[0].keys()))

    long_rows = build_long_core(fixed_total, fixed_snapshot, mobile, mobile_traffic, fixed_traffic)
    write_csv(OUTDIR / 'sector_core_monthly_long.csv', long_rows, list(long_rows[0].keys()))
    annual = [r for r in long_rows if r['month'] == 12 and r['year'] <= 2025]
    write_csv(OUTDIR / 'sector_core_december_long_2000_2025.csv', annual, list(annual[0].keys()))

    qa_rows = build_qa(fixed_total, fixed_snapshot, mobile, mobile_traffic, fixed_traffic)
    write_csv(OUTDIR / 'series_qa.csv', qa_rows, ['check', 'value', 'expectation'])
    q = {r['check']: str(r['value']) for r in qa_rows}

    for name in ['fixed_connections', 'mobile_connections', 'mobile_traffic', 'fixed_traffic']:
        if q[f'{name}_last_period'] != '2026-03':
            raise RuntimeError(f'{name} does not end in 2026-03')
    if int(float(q['2026m03_fixed_total'])) != 4_859_679:
        raise RuntimeError('Fixed total does not match March 2026 official workbook')
    if int(float(q['2026m03_fixed_snapshot_total'])) != 4_859_679:
        raise RuntimeError('Current fixed-technology snapshot does not reconcile to fixed total')
    if int(float(q['2026m03_fixed_fttx_connections'])) != 4_147_629:
        raise RuntimeError('Current explicit FTTX total changed')
    if not (85.2 <= float(q['2026m03_fixed_fttx_share_pct']) <= 85.5):
        raise RuntimeError('Current FTTX share outside expected range')
    if int(float(q['2026m01_5g_connections'])) != 10_161_957:
        raise RuntimeError('January 2026 5G value does not match official SUBTEL January publication')
    if int(float(q['2026m03_5g_connections_workbook'])) != 10_356_448:
        raise RuntimeError('March 2026 5G monthly workbook value changed')

    print('fixed_total', len(fixed_total), fixed_total[0]['period'], fixed_total[-1]['period'])
    print('fixed_snapshot', fixed_snapshot)
    print('mobile', len(mobile), mobile[0]['period'], mobile[-1]['period'])
    print('mobile_traffic', len(mobile_traffic), mobile_traffic[0]['period'], mobile_traffic[-1]['period'])
    print('fixed_traffic', len(fixed_traffic), fixed_traffic[0]['period'], fixed_traffic[-1]['period'])
    print('long_rows', len(long_rows), 'annual_december_rows', len(annual))
    print('5g_mar_workbook_minus_sector_snapshot', 10_356_448 - 10_367_754)


if __name__ == '__main__':
    main()
