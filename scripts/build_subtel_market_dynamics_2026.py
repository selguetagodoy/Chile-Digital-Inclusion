from __future__ import annotations

import csv
import io
import math
import re
from pathlib import Path

import openpyxl
import requests

OUTDIR = Path('data/subtel_sector_series')
LATEST_PERIOD = '2026-06'

SOURCES = {
    'fixed': 'https://www.subtel.gob.cl/wp-content/uploads/2026/08/1_SERIES_CONEXIONES_INTERNET_FIJA_JUN26_100826.xlsx',
    'mobile': 'https://www.subtel.gob.cl/wp-content/uploads/2026/08/2_SERIES_CONEXIONES_INTERNET_MO%CC%81VIL-JUN26-100826.xlsx',
    'mobile_traffic': 'https://www.subtel.gob.cl/wp-content/uploads/2026/08/3_SERIES_TRAFICO_DATOS_MOVILES-JUN26-100826.xlsx',
    'fixed_traffic': 'https://www.subtel.gob.cl/wp-content/uploads/2026/08/3_SERIES_TRAFICO_DATOS_FIJOS-JUN26-100826.xlsx',
}

MONTHS = {
    'ene': 1, 'feb': 2, 'mar': 3, 'abr': 4, 'may': 5, 'jun': 6,
    'jul': 7, 'ago': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dic': 12,
}


def download(url: str) -> bytes:
    r = requests.get(url, timeout=180, allow_redirects=True, headers={'User-Agent': 'Chile-Digital-Inclusion/2026'})
    r.raise_for_status()
    if not r.content.startswith(b'PK'):
        raise RuntimeError(f'Expected XLSX from {url}; content-type={r.headers.get("content-type")} bytes={len(r.content)}')
    return r.content


def workbook(key: str):
    return openpyxl.load_workbook(io.BytesIO(download(SOURCES[key])), read_only=True, data_only=True)


def number(v):
    if v is None or v == '' or v == '---':
        return None
    if isinstance(v, bool):
        return None
    try:
        x = float(v)
        return None if math.isnan(x) else x
    except (TypeError, ValueError):
        return None


def integer(v):
    x = number(v)
    return None if x is None else int(round(x))


def month_no(v):
    if v is None:
        return None
    return MONTHS.get(str(v).strip().lower().replace('.', '')[:3])


def period(year: int, month: int) -> str:
    return f'{year:04d}-{month:02d}'


def clean(v) -> str:
    return '' if v is None else str(v).strip()


def slug(v: str) -> str:
    s = v.lower().strip()
    replacements = {
        'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u', 'ü': 'u', 'ñ': 'n',
        '+': '_plus_', '&': '_and_'
    }
    for a, b in replacements.items():
        s = s.replace(a, b)
    s = re.sub(r'[^a-z0-9]+', '_', s).strip('_')
    return s


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open('w', encoding='utf-8', newline='') as fh:
        w = csv.DictWriter(fh, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)


def iter_rows(ws, start_row: int, year_col: int, month_col: int, max_col: int):
    current_year = None
    for row_no, row in enumerate(ws.iter_rows(min_row=start_row, max_col=max_col, values_only=True), start=start_row):
        vals = list(row)
        y = integer(vals[year_col - 1]) if year_col <= len(vals) else None
        if y is not None and 1990 <= y <= 2100:
            current_year = y
        m = month_no(vals[month_col - 1]) if month_col <= len(vals) else None
        if current_year is not None and m is not None:
            yield row_no, current_year, m, vals


def build_fixed_plan(wb):
    ws = wb['7.8.CO_PL_FIJAS']
    rows = []
    for row_no, y, m, v in iter_rows(ws, 7, 2, 3, 6):
        postpaid, prepaid, total = integer(v[3]), integer(v[4]), integer(v[5])
        if total is None:
            continue
        rows.append({
            'period': period(y, m), 'year': y, 'month': m,
            'postpaid_connections': postpaid,
            'prepaid_connections': prepaid,
            'total_connections': total,
            'postpaid_share_pct': round(postpaid / total * 100, 6) if postpaid is not None and total else None,
            'prepaid_share_pct': round(prepaid / total * 100, 6) if prepaid is not None and total else None,
            'source_sheet': '7.8.CO_PL_FIJAS', 'source_row': row_no,
        })
    return rows


def build_mobile_plan(wb):
    ws = wb['8.9.CO_MOVIL_PLAN']
    rows = []
    for row_no, y, m, v in iter_rows(ws, 9, 2, 3, 17):
        contract_3g5g = integer(v[7])
        contract_all = integer(v[8])
        prepaid_3g5g = integer(v[13])
        prepaid_all = integer(v[14])
        total_3g5g = integer(v[15])
        total_all = integer(v[16])
        if total_3g5g is None:
            continue
        rows.append({
            'period': period(y, m), 'year': y, 'month': m,
            'contract_3g_4g_5g_connections': contract_3g5g,
            'prepaid_3g_4g_5g_connections': prepaid_3g5g,
            'total_3g_4g_5g_connections': total_3g5g,
            'contract_2g_3g_4g_5g_connections': contract_all,
            'prepaid_2g_3g_4g_5g_connections': prepaid_all,
            'total_2g_3g_4g_5g_connections': total_all,
            'contract_share_3g_4g_5g_pct': round(contract_3g5g / total_3g5g * 100, 6) if contract_3g5g is not None and total_3g5g else None,
            'prepaid_share_3g_4g_5g_pct': round(prepaid_3g5g / total_3g5g * 100, 6) if prepaid_3g5g is not None and total_3g5g else None,
            'source_sheet': '8.9.CO_MOVIL_PLAN', 'source_row': row_no,
        })
    return rows


def nearest_year_month_block(headers: list, total_label: str):
    total_candidates = [i for i, h in enumerate(headers) if clean(h) == total_label]
    if not total_candidates:
        raise RuntimeError(f'Could not locate total label {total_label}')
    total_idx = total_candidates[-1]
    year_candidates = [i for i, h in enumerate(headers[:total_idx]) if clean(h) == 'Año']
    if not year_candidates:
        raise RuntimeError(f'Could not locate Año before {total_label}')
    year_idx = max(year_candidates)
    month_candidates = [i for i, h in enumerate(headers[year_idx + 1:total_idx], start=year_idx + 1) if clean(h) == 'Mes']
    if not month_candidates:
        raise RuntimeError(f'Could not locate Mes before {total_label}')
    month_idx = min(month_candidates)
    return year_idx, month_idx, total_idx


def build_operator_block(ws, header_row: int, data_start: int, total_label: str, source_sheet: str, value_name: str, aliases: dict[str, str] | None = None):
    headers = list(next(ws.iter_rows(min_row=header_row, max_row=header_row, max_col=ws.max_column, values_only=True)))
    year_idx, month_idx, total_idx = nearest_year_month_block(headers, total_label)
    operator_cols = [(i, clean(headers[i])) for i in range(month_idx + 1, total_idx) if clean(headers[i])]
    aliases = aliases or {}
    rows = []
    current_year = None
    for row_no, row in enumerate(ws.iter_rows(min_row=data_start, max_col=total_idx + 1, values_only=True), start=data_start):
        v = list(row)
        y = integer(v[year_idx]) if year_idx < len(v) else None
        if y is not None and 1990 <= y <= 2100:
            current_year = y
        m = month_no(v[month_idx]) if month_idx < len(v) else None
        total = number(v[total_idx]) if total_idx < len(v) else None
        if current_year is None or m is None or total is None:
            continue
        total_int = int(round(total)) if value_name == 'connections' else round(total, 6)
        for idx, source_operator in operator_cols:
            value = number(v[idx]) if idx < len(v) else None
            if value is None:
                continue
            value_out = int(round(value)) if value_name == 'connections' else round(value, 6)
            operator = aliases.get(source_operator, source_operator)
            rows.append({
                'period': period(current_year, m), 'year': current_year, 'month': m,
                'operator': operator, 'operator_source_label': source_operator,
                value_name: value_out,
                'total_market': total_int,
                'share_pct': round(value / total * 100, 6) if total else None,
                'source_sheet': source_sheet, 'source_row': row_no,
            })
    return rows


def build_fixed_operators(wb):
    aliases = {
        'Telefónica': 'Movistar',
        'Grupo ENTEL': 'Entel',
        'Grupo GTD': 'GTD',
    }
    return build_operator_block(
        wb['7.9.CO_EMPR_FIJAS'], 7, 8, 'Total general', '7.9.CO_EMPR_FIJAS', 'connections', aliases
    )


def build_mobile_operators(wb):
    aliases = {'Entel PCS': 'Entel'}
    return build_operator_block(
        wb['8.3.CO_EMP_TEC_MOVIL'], 8, 9, 'Total de Conexiones 3G+4G+5G', '8.3.CO_EMP_TEC_MOVIL', 'connections', aliases
    )


def build_satellite_operator_proxy(wb):
    ws = wb['7.9.CO_EMPR_FIJAS']
    headers = list(next(ws.iter_rows(min_row=7, max_row=7, max_col=ws.max_column, values_only=True)))
    first_year = next(i for i, h in enumerate(headers) if clean(h) == 'Año')
    first_month = next(i for i, h in enumerate(headers) if i > first_year and clean(h) == 'Mes')
    wanted = ['Hughesnet', 'Starlink']
    indices = {}
    for name in wanted:
        candidates = [i for i, h in enumerate(headers) if clean(h) == name]
        if not candidates:
            raise RuntimeError(f'Missing {name} in 7.9.CO_EMPR_FIJAS')
        indices[name] = candidates[0]
    total_candidates = [i for i, h in enumerate(headers) if clean(h).lower() == 'total de conexiones']
    total_idx = total_candidates[0] if total_candidates else None
    max_col = max([first_year, first_month, *indices.values(), total_idx or 0]) + 1
    rows = []
    current_year = None
    for row_no, row in enumerate(ws.iter_rows(min_row=8, max_col=max_col, values_only=True), start=8):
        v = list(row)
        y = integer(v[first_year]) if first_year < len(v) else None
        if y is not None and 1990 <= y <= 2100:
            current_year = y
        m = month_no(v[first_month]) if first_month < len(v) else None
        if current_year is None or m is None:
            continue
        vals = {name: integer(v[idx]) if idx < len(v) else None for name, idx in indices.items()}
        if all(x is None for x in vals.values()):
            continue
        total_market = integer(v[total_idx]) if total_idx is not None and total_idx < len(v) else None
        provider_sum = sum(x or 0 for x in vals.values())
        for name in wanted:
            value = vals[name]
            if value is None:
                continue
            rows.append({
                'period': period(current_year, m), 'year': current_year, 'month': m,
                'operator': name, 'connections': value,
                'two_named_satellite_operators_total': provider_sum,
                'total_fixed_market': total_market,
                'share_of_total_fixed_pct': round(value / total_market * 100, 6) if total_market else None,
                'interpretation': 'Operator-level satellite-provider series; not a complete technology-total measure of satellite access',
                'source_sheet': '7.9.CO_EMPR_FIJAS', 'source_row': row_no,
            })
    return rows


def build_mobile_traffic_operators(wb):
    ws = wb['9.5.TRAF_EMP']
    headers = list(next(ws.iter_rows(min_row=10, max_row=10, max_col=ws.max_column, values_only=True)))
    total_idx = next(i for i, h in enumerate(headers) if clean(h) == 'Total')
    operator_cols = [(i, clean(headers[i])) for i in range(3, total_idx) if clean(headers[i])]
    aliases = {'ENTEL': 'Entel'}
    rows = []
    current_year = None
    for row_no, row in enumerate(ws.iter_rows(min_row=11, max_col=total_idx + 1, values_only=True), start=11):
        v = list(row)
        y = integer(v[1])
        if y is not None and 1990 <= y <= 2100:
            current_year = y
        m = month_no(v[2])
        total = number(v[total_idx])
        if current_year is None or m is None or total is None:
            continue
        for idx, source_operator in operator_cols:
            value = number(v[idx])
            if value is None:
                continue
            rows.append({
                'period': period(current_year, m), 'year': current_year, 'month': m,
                'operator': aliases.get(source_operator, source_operator), 'operator_source_label': source_operator,
                'traffic_tb': round(value, 6), 'total_market_tb': round(total, 6),
                'share_pct': round(value / total * 100, 6) if total else None,
                'source_sheet': '9.5.TRAF_EMP', 'source_row': row_no,
            })
    return rows


def build_fixed_traffic_operators(wb):
    ws = wb['10.4.TRAF_EMP']
    headers = list(next(ws.iter_rows(min_row=10, max_row=10, max_col=ws.max_column, values_only=True)))
    total_idx = next(i for i, h in enumerate(headers) if clean(h) == 'Total')
    operator_cols = [(i, clean(headers[i])) for i in range(3, total_idx) if clean(headers[i])]
    aliases = {'Entel S.A.': 'Entel', 'Grupo Claro-VTR': 'Grupo Claro-VTR'}
    rows = []
    current_year = None
    for row_no, row in enumerate(ws.iter_rows(min_row=11, max_col=total_idx + 1, values_only=True), start=11):
        v = list(row)
        y = integer(v[1])
        if y is not None and 1990 <= y <= 2100:
            current_year = y
        m = month_no(v[2])
        total = number(v[total_idx])
        if current_year is None or m is None or total is None:
            continue
        for idx, source_operator in operator_cols:
            value = number(v[idx])
            if value is None:
                continue
            rows.append({
                'period': period(current_year, m), 'year': current_year, 'month': m,
                'operator': aliases.get(source_operator, source_operator), 'operator_source_label': source_operator,
                'traffic_tb': round(value, 6), 'total_market_tb': round(total, 6),
                'share_pct': round(value / total * 100, 6) if total else None,
                'source_sheet': '10.4.TRAF_EMP', 'source_row': row_no,
            })
    return rows


def build_mobile_traffic_plan(wb):
    ws = wb['9.3.TRAF_CLI.PLAN']
    rows = []
    for row_no, y, m, v in iter_rows(ws, 12, 2, 3, 7):
        residential_postpaid = number(v[3])
        commercial_postpaid = number(v[4])
        prepaid = number(v[5])
        total = number(v[6])
        if total is None:
            continue
        postpaid = (residential_postpaid or 0) + (commercial_postpaid or 0)
        rows.append({
            'period': period(y, m), 'year': y, 'month': m,
            'residential_postpaid_traffic_tb': round(residential_postpaid, 6) if residential_postpaid is not None else None,
            'commercial_postpaid_traffic_tb': round(commercial_postpaid, 6) if commercial_postpaid is not None else None,
            'postpaid_traffic_tb': round(postpaid, 6),
            'prepaid_traffic_tb': round(prepaid, 6) if prepaid is not None else None,
            'total_traffic_tb': round(total, 6),
            'prepaid_traffic_share_pct': round(prepaid / total * 100, 6) if prepaid is not None and total else None,
            'postpaid_traffic_share_pct': round(postpaid / total * 100, 6) if total else None,
            'source_sheet': '9.3.TRAF_CLI.PLAN', 'source_row': row_no,
        })
    return rows


def comparison_rows(domain: str, metric: str, data: list[dict], entity_key: str, value_key: str, periods=(('2024-06', '2025-06'), ('2025-06', '2026-06'), ('2025-12', '2026-06'))):
    idx = {(r['period'], str(r[entity_key])): r for r in data}
    entities = sorted({str(r[entity_key]) for r in data})
    out = []
    for base_period, end_period in periods:
        for entity in entities:
            a = idx.get((base_period, entity))
            b = idx.get((end_period, entity))
            if not a or not b:
                continue
            av, bv = number(a.get(value_key)), number(b.get(value_key))
            if av is None or bv is None:
                continue
            out.append({
                'domain': domain, 'metric': metric, 'entity': entity,
                'base_period': base_period, 'end_period': end_period,
                'base_value': round(av, 6), 'end_value': round(bv, 6),
                'absolute_change': round(bv - av, 6),
                'pct_change': round((bv / av - 1) * 100, 6) if av != 0 else None,
                'base_share_pct': a.get('share_pct'), 'end_share_pct': b.get('share_pct'),
                'share_change_pp': round(number(b.get('share_pct')) - number(a.get('share_pct')), 6) if number(a.get('share_pct')) is not None and number(b.get('share_pct')) is not None else None,
            })
    return out


def plan_comparisons(fixed_plan, mobile_plan, mobile_traffic_plan):
    out = []
    periods = [('2024-06', '2025-06'), ('2025-06', '2026-06'), ('2025-12', '2026-06')]
    specs = [
        ('fixed_plan', 'connections', fixed_plan, 'postpaid', 'postpaid_connections', 'postpaid_share_pct'),
        ('fixed_plan', 'connections', fixed_plan, 'prepaid', 'prepaid_connections', 'prepaid_share_pct'),
        ('mobile_plan', '3g_4g_5g_connections', mobile_plan, 'contract', 'contract_3g_4g_5g_connections', 'contract_share_3g_4g_5g_pct'),
        ('mobile_plan', '3g_4g_5g_connections', mobile_plan, 'prepaid', 'prepaid_3g_4g_5g_connections', 'prepaid_share_3g_4g_5g_pct'),
        ('mobile_traffic_plan', 'traffic_tb', mobile_traffic_plan, 'postpaid', 'postpaid_traffic_tb', 'postpaid_traffic_share_pct'),
        ('mobile_traffic_plan', 'traffic_tb', mobile_traffic_plan, 'prepaid', 'prepaid_traffic_tb', 'prepaid_traffic_share_pct'),
    ]
    for domain, metric, data, entity, value_key, share_key in specs:
        by_period = {r['period']: r for r in data}
        for base_period, end_period in periods:
            a, b = by_period.get(base_period), by_period.get(end_period)
            if not a or not b:
                continue
            av, bv = number(a.get(value_key)), number(b.get(value_key))
            if av is None or bv is None:
                continue
            ash, bsh = number(a.get(share_key)), number(b.get(share_key))
            out.append({
                'domain': domain, 'metric': metric, 'entity': entity,
                'base_period': base_period, 'end_period': end_period,
                'base_value': round(av, 6), 'end_value': round(bv, 6),
                'absolute_change': round(bv - av, 6),
                'pct_change': round((bv / av - 1) * 100, 6) if av != 0 else None,
                'base_share_pct': ash, 'end_share_pct': bsh,
                'share_change_pp': round(bsh - ash, 6) if ash is not None and bsh is not None else None,
            })
    return out


def validate_latest(rows: list[dict], name: str):
    periods = [r['period'] for r in rows]
    if not periods or max(periods) != LATEST_PERIOD:
        raise RuntimeError(f'{name} does not reach {LATEST_PERIOD}; last={max(periods) if periods else None}')


def main():
    OUTDIR.mkdir(parents=True, exist_ok=True)
    fixed_wb = workbook('fixed')
    mobile_wb = workbook('mobile')
    mobile_traffic_wb = workbook('mobile_traffic')
    fixed_traffic_wb = workbook('fixed_traffic')

    fixed_plan = build_fixed_plan(fixed_wb)
    mobile_plan = build_mobile_plan(mobile_wb)
    fixed_ops = build_fixed_operators(fixed_wb)
    mobile_ops = build_mobile_operators(mobile_wb)
    satellite = build_satellite_operator_proxy(fixed_wb)
    mobile_traffic_ops = build_mobile_traffic_operators(mobile_traffic_wb)
    fixed_traffic_ops = build_fixed_traffic_operators(fixed_traffic_wb)
    mobile_traffic_plan = build_mobile_traffic_plan(mobile_traffic_wb)

    datasets = {
        'fixed_connections_by_plan_monthly.csv': fixed_plan,
        'mobile_connections_by_plan_monthly.csv': mobile_plan,
        'fixed_connections_by_operator_monthly.csv': fixed_ops,
        'mobile_connections_by_operator_monthly.csv': mobile_ops,
        'fixed_satellite_operator_connections_monthly.csv': satellite,
        'mobile_traffic_by_operator_monthly.csv': mobile_traffic_ops,
        'fixed_traffic_by_operator_monthly.csv': fixed_traffic_ops,
        'mobile_traffic_by_plan_monthly.csv': mobile_traffic_plan,
    }
    for filename, rows in datasets.items():
        validate_latest(rows, filename)
        write_csv(OUTDIR / filename, rows, list(rows[0].keys()))

    comparisons = []
    comparisons.extend(plan_comparisons(fixed_plan, mobile_plan, mobile_traffic_plan))
    comparisons.extend(comparison_rows('fixed_operator', 'connections', fixed_ops, 'operator', 'connections'))
    comparisons.extend(comparison_rows('mobile_operator_3g_4g_5g', 'connections', mobile_ops, 'operator', 'connections'))
    comparisons.extend(comparison_rows('satellite_operator_proxy', 'connections', satellite, 'operator', 'connections'))
    comparisons.extend(comparison_rows('mobile_traffic_operator', 'traffic_tb', mobile_traffic_ops, 'operator', 'traffic_tb'))
    comparisons.extend(comparison_rows('fixed_traffic_operator', 'traffic_tb', fixed_traffic_ops, 'operator', 'traffic_tb'))
    write_csv(OUTDIR / 'market_dynamics_comparisons_2024_2026.csv', comparisons, [
        'domain', 'metric', 'entity', 'base_period', 'end_period', 'base_value', 'end_value',
        'absolute_change', 'pct_change', 'base_share_pct', 'end_share_pct', 'share_change_pp'
    ])

    qa = []
    for filename, rows in datasets.items():
        qa.append({'dataset': filename, 'rows': len(rows), 'first_period': min(r['period'] for r in rows), 'last_period': max(r['period'] for r in rows)})
    write_csv(OUTDIR / 'market_dynamics_qa.csv', qa, ['dataset', 'rows', 'first_period', 'last_period'])

    for r in qa:
        print(r)
    print('comparison_rows', len(comparisons))


if __name__ == '__main__':
    main()
