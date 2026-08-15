#!/usr/bin/env python3
"""Build March-2026 fixed Internet connections by commune from official SUBTEL XLSX.

The workbook's commune labels are stale/misaligned after the Biobío/Ñuble split,
but March-2026 values remain organized in 16 formula-defined regional blocks.
This builder maps rows inside each regional block to the current official commune
catalogue in normalized alphabetical order and requires exact count and subtotal
reconciliation for both total and residential connections. An explicit blank in
a one-row-per-commune block is preserved as source_blank rather than imputed zero.
"""
from __future__ import annotations

import csv
import io
import re
import unicodedata
import urllib.request
from pathlib import Path

import openpyxl

URL = "https://www.subtel.gob.cl/wp-content/uploads/2026/05/1_SERIES_CONEXIONES_INTERNET_FIJA_MAR26_040526.xlsx"
SHEETS = {
    "total_fixed_connections_2026m03": "7.11.CO_FIJAS_COMUNA",
    "residential_fixed_connections_2026m03": "7.11.1.CO_FIJAS_RES_COMUNA",
}
COMMUNES = Path("geo/commune_codes.csv")
OUT = Path("data/fixed_infrastructure_2026")
OUT.mkdir(parents=True, exist_ok=True)
ALIASES = {"calera": "la calera", "aisen": "aysen", "coihaique": "coyhaique", "treguaco": "trehuaco", "til til": "tiltil"}
REGION_CODES = list(range(1, 17))
EXPECTED_SOURCE_BLANK_CODES = {12202}


def norm(s):
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode("ascii").lower().strip()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return ALIASES.get(s, s)


def clean(v):
    return "" if v is None else str(v).strip().replace("\n", " ")


def as_int(raw):
    try:
        return int(round(float(raw)))
    except (TypeError, ValueError):
        return None


def col_letters(n):
    out = ""
    while n:
        n, rem = divmod(n - 1, 26)
        out = chr(65 + rem) + out
    return out


def locate_period_column(ws):
    current_year = None
    matches = []
    for col in range(1, ws.max_column + 1):
        y = ws.cell(8, col).value
        if isinstance(y, (int, float)) and int(y) == y and 2000 <= int(y) <= 2100:
            current_year = int(y)
        month = clean(ws.cell(9, col).value).lower()
        if current_year == 2026 and month in {"mar", "marzo"}:
            matches.append(col)
    if not matches:
        raise RuntimeError("Could not locate March 2026 column")
    return max(matches)


def regional_formula_blocks(ws_values, ws_formulas, target_col):
    letters = col_letters(target_col)
    pattern = re.compile(rf"^=?SUM\(\$?{letters}\$?(\d+):\$?{letters}\$?(\d+)\)$", re.I)
    blocks = []
    for rn in range(10, ws_formulas.max_row + 1):
        formula = clean(ws_formulas.cell(rn, target_col).value).replace(" ", "")
        m = pattern.match(formula)
        if not m:
            continue
        blocks.append(
            {
                "subtotal_row": rn,
                "start_row": int(m.group(1)),
                "end_row": int(m.group(2)),
                "subtotal": as_int(ws_values.cell(rn, target_col).value),
            }
        )
    if len(blocks) != 16:
        raise RuntimeError(f"Expected 16 regional subtotal formula blocks, found {len(blocks)}")
    for region, block in zip(REGION_CODES, blocks):
        block["region"] = region
    return blocks


def build_metric(ws_values, ws_formulas, target_col, catalogue_by_region, metric):
    mapped = {}
    alignment = []
    provenance = []
    issues = []

    for block in regional_formula_blocks(ws_values, ws_formulas, target_col):
        region = block["region"]
        communes = catalogue_by_region[region]
        block_slots = []
        for rn in range(block["start_row"], block["end_row"] + 1):
            formula = clean(ws_formulas.cell(rn, target_col).value)
            source_label = clean(ws_values.cell(rn, 3).value)
            value = None if formula.startswith("=") else as_int(ws_values.cell(rn, target_col).value)
            block_slots.append((rn, source_label, value))

        # When the formula range has exactly one row per current commune, preserve
        # blank cells positionally. Otherwise remove non-numeric auxiliary rows
        # such as historical "Sin clasificación" rows and require the remaining
        # numeric observations to equal the current commune count.
        if len(block_slots) == len(communes):
            observations = block_slots
        else:
            observations = [slot for slot in block_slots if slot[2] is not None]

        count_ok = len(observations) == len(communes)
        value_sum = sum(v for _, _, v in observations if v is not None)
        subtotal = block["subtotal"]
        subtotal_ok = subtotal is not None and value_sum == subtotal
        label_matches = 0
        blank_slots = 0

        if count_ok:
            for commune, (rn, source_label, value) in zip(communes, observations):
                code = int(commune["comuna"])
                if value is not None:
                    mapped[code] = value
                else:
                    blank_slots += 1
                if norm(source_label) == norm(commune["comuna_nombre"]):
                    label_matches += 1
                provenance.append(
                    {
                        "metric": metric,
                        "region": region,
                        "source_row": rn,
                        "source_commune_label": source_label,
                        "mapped_commune": code,
                        "mapped_commune_name": commune["comuna_nombre"],
                        "value": "" if value is None else value,
                        "source_cell_status": "blank" if value is None else "numeric",
                        "mapping_method": "regional_formula_block_order",
                        "block_start_row": block["start_row"],
                        "block_end_row": block["end_row"],
                        "regional_subtotal_row": block["subtotal_row"],
                    }
                )

        alignment.append(
            {
                "metric": metric,
                "region": region,
                "block_start_row": block["start_row"],
                "block_end_row": block["end_row"],
                "block_rows": len(block_slots),
                "regional_subtotal_row": block["subtotal_row"],
                "mapped_commune_slots": len(observations),
                "catalogue_communes": len(communes),
                "source_blank_slots": blank_slots,
                "regional_subtotal": subtotal,
                "mapped_value_sum": value_sum,
                "subtotal_delta": "" if subtotal is None else value_sum - subtotal,
                "source_labels_matching_mapped_names": label_matches,
                "count_status": "pass" if count_ok else "fail",
                "subtotal_status": "pass" if subtotal_ok else "fail",
            }
        )
        if not count_ok or not subtotal_ok:
            issues.append(
                [metric, region, block["start_row"], block["end_row"], len(observations), len(communes), subtotal, value_sum]
            )

    return mapped, alignment, provenance, issues


def write_dict_csv(path, fields, rows):
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)


def main():
    request = urllib.request.Request(URL, headers={"User-Agent": "Mozilla/5.0 Chile-Digital-Inclusion/1.0"})
    with urllib.request.urlopen(request, timeout=120) as response:
        content = response.read()

    wb_values = openpyxl.load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    wb_formulas = openpyxl.load_workbook(io.BytesIO(content), read_only=False, data_only=False)

    with COMMUNES.open(encoding="utf-8") as f:
        catalogue = list(csv.DictReader(f))
    catalogue_by_region = {}
    for region in REGION_CODES:
        region_rows = [r for r in catalogue if int(r["region"]) == region]
        catalogue_by_region[region] = sorted(region_rows, key=lambda r: norm(r["comuna_nombre"]))

    all_alignment = []
    all_provenance = []
    all_issues = []
    metric_maps = {}
    cols = {}

    for metric, sheet_name in SHEETS.items():
        if sheet_name not in wb_values.sheetnames:
            raise RuntimeError(f"Expected sheet not found: {sheet_name}")
        ws_values = wb_values[sheet_name]
        ws_formulas = wb_formulas[sheet_name]
        target_col = locate_period_column(ws_values)
        cols[metric] = target_col
        mapped, alignment, provenance, issues = build_metric(
            ws_values, ws_formulas, target_col, catalogue_by_region, metric
        )
        metric_maps[metric] = mapped
        all_alignment.extend(alignment)
        all_provenance.extend(provenance)
        all_issues.extend(issues)

    if all_issues:
        raise RuntimeError(f"Regional formula-block reconciliation failed: {all_issues}")

    by_code = {int(r["comuna"]): r for r in catalogue}
    output_rows = []
    missing = []
    for code in sorted(by_code):
        r = by_code[code]
        total = metric_maps["total_fixed_connections_2026m03"].get(code)
        residential = metric_maps["residential_fixed_connections_2026m03"].get(code)
        share = round(residential / total * 100, 4) if total and residential is not None else None
        status = "reported" if total is not None and residential is not None else "source_blank"
        output_rows.append(
            [
                r["region"], r["region_nombre"], r["provincia"], r["provincia_nombre"], r["comuna"], r["comuna_nombre"],
                total, residential, share, status, "regional_formula_block_order", "2026-03", URL,
            ]
        )
        if status != "reported":
            missing.append([r["region"], r["region_nombre"], r["comuna"], r["comuna_nombre"], total, residential, status])

    with (OUT / "commune_fixed_connections_2026_03.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([
            "region", "region_nombre", "provincia", "provincia_nombre", "comuna", "comuna_nombre",
            "fixed_connections_total", "fixed_connections_residential", "residential_share_pct", "source_status",
            "source_mapping_method", "period", "source_url",
        ])
        w.writerows(output_rows)

    with (OUT / "source_match_qa.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["metric", "region", "block_start_row", "block_end_row", "mapped_slots", "catalogue_communes", "regional_subtotal", "mapped_value_sum"])
        w.writerows(all_issues)

    with (OUT / "source_not_reported_communes.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["region", "region_nombre", "comuna", "comuna_nombre", "fixed_connections_total", "fixed_connections_residential", "source_status"])
        w.writerows(missing)

    alignment_fields = [
        "metric", "region", "block_start_row", "block_end_row", "block_rows", "regional_subtotal_row",
        "mapped_commune_slots", "catalogue_communes", "source_blank_slots", "regional_subtotal", "mapped_value_sum",
        "subtotal_delta", "source_labels_matching_mapped_names", "count_status", "subtotal_status",
    ]
    write_dict_csv(OUT / "source_alignment_qa.csv", alignment_fields, all_alignment)

    provenance_fields = [
        "metric", "region", "source_row", "source_commune_label", "mapped_commune", "mapped_commune_name", "value",
        "source_cell_status", "mapping_method", "block_start_row", "block_end_row", "regional_subtotal_row",
    ]
    write_dict_csv(OUT / "source_row_mapping_2026_03.csv", provenance_fields, all_provenance)

    with (OUT / "extraction_manifest.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["metric", "source_sheet", "march_2026_column_number", "mapping_method", "source_url"])
        for metric, sheet_name in SHEETS.items():
            w.writerow([metric, sheet_name, cols[metric], "regional_formula_block_order_with_subtotal_reconciliation", URL])

    total_count = len(metric_maps["total_fixed_connections_2026m03"])
    residential_count = len(metric_maps["residential_fixed_connections_2026m03"])
    missing_codes = {int(r[2]) for r in missing}
    print(f"communes total fixed numeric: {total_count}/346")
    print(f"communes residential fixed numeric: {residential_count}/346")
    print(f"explicit source blanks: {len(missing)} {sorted(missing_codes)}")
    print(f"regional alignment checks: {len(all_alignment)} all pass")
    if total_count != 345 or residential_count != 345 or missing_codes != EXPECTED_SOURCE_BLANK_CODES:
        raise SystemExit("Unexpected formula-block reconstruction coverage")


if __name__ == "__main__":
    main()
