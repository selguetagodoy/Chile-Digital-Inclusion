from __future__ import annotations

import csv
import io
from pathlib import Path

import openpyxl
import requests

URL='https://www.subtel.gob.cl/wp-content/uploads/2026/05/1_SERIES_CONEXIONES_INTERNET_FIJA_MAR26_040526.xlsx'
OUT=Path('data/subtel_sector_series/current_fixed_technology_sheet_audit.csv')
SHEETS=['7.7.1.CO_TEC_RG_EMP_FIJAS','7.13.CO_FIJAS_TECN_OECD']


def clean(v):
    return '' if v is None else str(v).strip().replace('\n',' ')


def main():
    r=requests.get(URL,timeout=180); r.raise_for_status()
    wb=openpyxl.load_workbook(io.BytesIO(r.content),read_only=True,data_only=True)
    rows=[]
    for name in SHEETS:
        ws=wb[name]
        # First 18 non-empty rows and final 8 non-empty rows, up to 40 columns.
        nonempty=[]
        for row_no,row in enumerate(ws.iter_rows(min_row=1,max_row=ws.max_row,max_col=min(ws.max_column,40),values_only=True),start=1):
            vals=[clean(v) for v in row]
            if any(vals):
                nonempty.append((row_no,vals))
        selected=nonempty[:18]+nonempty[-8:]
        seen=set()
        for row_no,vals in selected:
            if row_no in seen: continue
            seen.add(row_no)
            rows.append({'sheet':name,'source_row':row_no,'values':' || '.join(vals)})
    OUT.parent.mkdir(parents=True,exist_ok=True)
    with OUT.open('w',encoding='utf-8',newline='') as fh:
        w=csv.DictWriter(fh,fieldnames=['sheet','source_row','values']); w.writeheader(); w.writerows(rows)
    print('audit_rows',len(rows))
    for r in rows: print(r)

if __name__=='__main__': main()
