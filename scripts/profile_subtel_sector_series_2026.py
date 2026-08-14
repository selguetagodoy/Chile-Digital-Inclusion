from __future__ import annotations

import csv
import io
import re
from pathlib import Path

import openpyxl
import requests

OUT = Path('data/subtel_sector_series/official_workbook_profile.csv')
HEADERS = Path('data/subtel_sector_series/official_workbook_headers.csv')

SOURCES = [
    ('fixed_connections', 'https://www.subtel.gob.cl/wp-content/uploads/2026/05/1_SERIES_CONEXIONES_INTERNET_FIJA_MAR26_040526.xlsx'),
    ('mobile_connections', 'https://www.subtel.gob.cl/wp-content/uploads/2026/05/2_SERIES_CONEXIONES_INTERNET_MO%CC%81VIL-MAR26-040526.xlsx'),
    ('mobile_traffic', 'https://www.subtel.gob.cl/wp-content/uploads/2026/05/3_SERIES_TRAFICO_DATOS_MOVILES-MAR26-040526.xlsx'),
    ('fixed_traffic', 'https://www.subtel.gob.cl/wp-content/uploads/2026/05/3_SERIES_TRAFICO_DATOS_FIJOS-MAR26-040526.xlsx'),
]


def clean(v):
    return re.sub(r'\s+', ' ', '' if v is None else str(v)).strip()


def score_row(values):
    joined = ' | '.join(clean(v).lower() for v in values if clean(v))
    terms = ['año', 'ano', 'mes', 'trimestre', 'total', 'conex', 'tráfico', 'trafico', 'tecnolog', 'operador', 'empresa', '4g', '5g', 'fibra']
    return sum(t in joined for t in terms)


def download(url: str) -> bytes:
    r = requests.get(url, timeout=180, allow_redirects=True)
    r.raise_for_status()
    body = r.content
    if not body.startswith(b'PK'):
        raise RuntimeError(f'Unexpected workbook payload from {url}: type={r.headers.get("content-type")} bytes={len(body)}')
    return body


def main():
    OUT.parent.mkdir(parents=True, exist_ok=True)
    profiles = []
    headers = []

    for dataset, url in SOURCES:
        body = download(url)
        wb = openpyxl.load_workbook(io.BytesIO(body), read_only=True, data_only=True)
        for ws in wb.worksheets:
            preview = []
            for row_no, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 40), values_only=True), start=1):
                values = list(row)
                preview.append((row_no, values, score_row(values)))
            best = max(preview, key=lambda x: x[2]) if preview else (1, [], 0)
            profiles.append({
                'dataset': dataset,
                'source_url': url,
                'workbook_bytes': len(body),
                'sheet': ws.title,
                'max_row': ws.max_row,
                'max_column': ws.max_column,
                'candidate_header_row': best[0],
                'header_score': best[2],
            })
            for col_no, value in enumerate(best[1], start=1):
                if clean(value):
                    headers.append({
                        'dataset': dataset,
                        'sheet': ws.title,
                        'header_row': best[0],
                        'column': col_no,
                        'header': clean(value),
                        'source_url': url,
                    })

    with OUT.open('w', encoding='utf-8', newline='') as fh:
        fields = list(profiles[0].keys())
        w = csv.DictWriter(fh, fieldnames=fields); w.writeheader(); w.writerows(profiles)
    with HEADERS.open('w', encoding='utf-8', newline='') as fh:
        fields = ['dataset','sheet','header_row','column','header','source_url']
        w = csv.DictWriter(fh, fieldnames=fields); w.writeheader(); w.writerows(headers)

    print('profile_rows', len(profiles), 'header_cells', len(headers))
    for row in profiles:
        print(row)


if __name__ == '__main__':
    main()
