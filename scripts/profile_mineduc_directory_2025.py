from __future__ import annotations

import csv
import subprocess
import tempfile
from pathlib import Path

import openpyxl
import requests

SOURCE_URL='https://datosabiertos.mineduc.cl/wp-content/uploads/2025/11/Directorio-Oficial-EE-2025.rar'
OUT=Path('data/education_connectivity_2026/mineduc_directory_2025_profile.csv')
HEADERS=Path('data/education_connectivity_2026/mineduc_directory_2025_headers.csv')


def download() -> bytes:
    r=requests.get(SOURCE_URL,timeout=180,allow_redirects=True)
    r.raise_for_status()
    if len(r.content) < 10000:
        raise RuntimeError(f'Archive unexpectedly small: {len(r.content)} bytes')
    return r.content


def clean(v):
    return '' if v is None else str(v).strip()


def profile_csv(path: Path):
    encodings=['utf-8-sig','latin1','cp1252']
    last=None
    for enc in encodings:
        try:
            with path.open(encoding=enc,newline='') as fh:
                sample=fh.read(4096); fh.seek(0)
                try:
                    dialect=csv.Sniffer().sniff(sample,delimiters=';,|\t,')
                except csv.Error:
                    dialect=csv.excel
                reader=csv.reader(fh,dialect)
                first=next(reader)
                rows=1+sum(1 for _ in reader)
            return first,rows,enc
        except Exception as exc:
            last=exc
    raise last


def extract_archive(archive: Path, extract: Path) -> str:
    attempts=[]
    commands=[
        ['unar','-quiet','-force-overwrite','-output-directory',str(extract),str(archive)],
        ['7z','x','-y',f'-o{extract}',str(archive)],
    ]
    for cmd in commands:
        proc=subprocess.run(cmd,capture_output=True,text=True)
        attempts.append(f"{' '.join(cmd[:1])}: rc={proc.returncode} stdout={proc.stdout[-300:]} stderr={proc.stderr[-300:]}")
        if proc.returncode == 0 and any(p.is_file() for p in extract.rglob('*')):
            return cmd[0]
    raise RuntimeError('Archive extraction failed: '+' | '.join(attempts))


def main():
    OUT.parent.mkdir(parents=True,exist_ok=True)
    body=download(); profiles=[]; headers=[]
    with tempfile.TemporaryDirectory() as td:
        root=Path(td); archive=root/'directory.rar'; archive.write_bytes(body)
        extract=root/'extract'; extract.mkdir()
        extractor=extract_archive(archive,extract)
        files=[p for p in extract.rglob('*') if p.is_file()]
        if not files:
            raise RuntimeError('Archive extracted no files')
        for path in files:
            suffix=path.suffix.lower()
            profile={'archive_bytes':len(body),'extractor':extractor,'file_name':path.name,'suffix':suffix,'file_bytes':path.stat().st_size,'rows':'','columns':'','sheet':'','encoding':'','source_url':SOURCE_URL}
            header=[]
            if suffix in {'.xlsx','.xlsm'}:
                wb=openpyxl.load_workbook(path,read_only=True,data_only=True)
                for ws in wb.worksheets:
                    vals=next(ws.iter_rows(min_row=1,max_row=1,values_only=True),())
                    header=[clean(v) for v in vals]
                    local=profile.copy(); local['rows']=ws.max_row; local['columns']=ws.max_column; local['sheet']=ws.title
                    profiles.append(local)
                    for i,h in enumerate(header,start=1):
                        if h: headers.append({'file_name':path.name,'sheet':ws.title,'column':i,'header':h})
                continue
            elif suffix in {'.csv','.txt'}:
                header,rows,enc=profile_csv(path)
                profile['rows']=rows; profile['columns']=len(header); profile['encoding']=enc
            profiles.append(profile)
            for i,h in enumerate(header,start=1):
                if clean(h): headers.append({'file_name':path.name,'sheet':'','column':i,'header':clean(h)})

    with OUT.open('w',encoding='utf-8',newline='') as fh:
        fields=list(profiles[0].keys()); w=csv.DictWriter(fh,fieldnames=fields); w.writeheader(); w.writerows(profiles)
    with HEADERS.open('w',encoding='utf-8',newline='') as fh:
        fields=['file_name','sheet','column','header']; w=csv.DictWriter(fh,fieldnames=fields); w.writeheader(); w.writerows(headers)
    print('profiles',profiles)
    print('headers',len(headers))

if __name__=='__main__':
    main()
