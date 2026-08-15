from __future__ import annotations

import csv
import io
from pathlib import Path

import openpyxl
import requests

URL = "https://www.subtel.gob.cl/wp-content/uploads/2026/05/1_SERIES_CONEXIONES_INTERNET_FIJA_MAR26_040526.xlsx"
OUT = Path("data/subtel_sector_series/fixed_commune_sheet_audit.csv")
DETAIL_OUT = Path("data/subtel_sector_series/fixed_commune_source_rows_2026_03.csv")
SHEETS = {
    "total": "7.11.CO_FIJAS_COMUNA",
    "residential": "7.11.1.CO_FIJAS_RES_COMUNA",
}
YEAR = 2026
MONTH_NAMES = {"mar", "marzo"}


def clean(v):
    return "" if v is None else str(v).strip().replace("\n", " ")


def locate_period_column(ws):
    current_year = None
    matches = []
    for col in range(1, ws.max_column + 1):
        y = ws.cell(8, col).value
        if isinstance(y, (int, float)) and int(y) == y and 2000 <= int(y) <= 2100:
            current_year = int(y)
        month = clean(ws.cell(9, col).value).lower()
        if current_year == YEAR and month in MONTH_NAMES:
            matches.append(col)
    if not matches:
        raise RuntimeError("Could not locate March 2026 column")
    return max(matches)


def formula_value(cell):
    value = cell.value
    return value if isinstance(value, str) and value.startswith("=") else ""


def main():
    r = requests.get(URL, timeout=180)
    r.raise_for_status()

    wb_values = openpyxl.load_workbook(io.BytesIO(r.content), read_only=False, data_only=True)
    wb_formulas = openpyxl.load_workbook(io.BytesIO(r.content), read_only=False, data_only=False)

    summary_rows = []
    detail_rows = []

    for metric, sheet_name in SHEETS.items():
        if sheet_name not in wb_values.sheetnames:
            raise RuntimeError(f"Expected sheet not found: {sheet_name}")

        ws = wb_values[sheet_name]
        wsf = wb_formulas[sheet_name]
        target_col = locate_period_column(ws)
        formula_rows = []
        suspicious_formula_labels = []
        carried_region = ""

        for row_no in range(10, ws.max_row + 1):
            region_label = clean(ws.cell(row_no, 2).value)
            commune_label = clean(ws.cell(row_no, 3).value)
            cached_value = clean(ws.cell(row_no, target_col).value)
            formula = formula_value(wsf.cell(row_no, target_col))

            if region_label:
                try:
                    carried_region = str(int(float(region_label)))
                except ValueError:
                    pass

            is_formula = bool(formula)
            if is_formula:
                item = (row_no, region_label, commune_label, cached_value, formula)
                formula_rows.append(item)
                if commune_label:
                    suspicious_formula_labels.append(item)

            if region_label or commune_label or cached_value or formula:
                detail_rows.append(
                    {
                        "metric": metric,
                        "source_sheet": sheet_name,
                        "source_row": row_no,
                        "region_label": region_label,
                        "carried_region": carried_region,
                        "commune_label": commune_label,
                        "march_2026_value": cached_value,
                        "march_2026_formula": formula,
                        "is_formula": "yes" if is_formula else "no",
                    }
                )

        summary_rows.extend(
            [
                {
                    "check": f"{metric}_sheet_present",
                    "status": "pass",
                    "value": sheet_name,
                    "detail": "Official March-2026 fixed Internet workbook commune sheet.",
                },
                {
                    "check": f"{metric}_march_2026_column",
                    "status": "pass",
                    "value": str(target_col),
                    "detail": f"Column located dynamically from year/month headers ({YEAR}-03).",
                },
                {
                    "check": f"{metric}_formula_rows_in_target_column",
                    "status": "info",
                    "value": str(len(formula_rows)),
                    "detail": "Subtotal/total formulas found in the March-2026 column.",
                },
                {
                    "check": f"{metric}_formula_rows_with_commune_label",
                    "status": "source_warning" if suspicious_formula_labels else "pass",
                    "value": str(len(suspicious_formula_labels)),
                    "detail": "Formula subtotal/total rows carrying commune labels indicate row-label/value misalignment in the official workbook; these rows must not be treated as commune observations.",
                },
            ]
        )

        for row_no, region_label, commune_label, cached_value, formula in suspicious_formula_labels:
            summary_rows.append(
                {
                    "check": f"{metric}_source_row_{row_no}",
                    "status": "source_warning",
                    "value": cached_value,
                    "detail": f"region_label={region_label or '<blank>'}; commune_label={commune_label}; formula={formula}",
                }
            )

        print(metric, sheet_name, "march_col", target_col, "formula_rows", len(formula_rows), "formula_labels", len(suspicious_formula_labels))

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=["check", "status", "value", "detail"])
        w.writeheader()
        w.writerows(summary_rows)

    with DETAIL_OUT.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(
            fh,
            fieldnames=[
                "metric",
                "source_sheet",
                "source_row",
                "region_label",
                "carried_region",
                "commune_label",
                "march_2026_value",
                "march_2026_formula",
                "is_formula",
            ],
        )
        w.writeheader()
        w.writerows(detail_rows)

    print("detail_rows", len(detail_rows))
    print("detail_output", DETAIL_OUT)


if __name__ == "__main__":
    main()
