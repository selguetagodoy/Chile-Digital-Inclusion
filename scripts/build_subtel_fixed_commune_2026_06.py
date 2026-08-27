#!/usr/bin/env python3
"""Build June-2026 fixed Internet connections by commune from official SUBTEL XLSX.

This reuses the validated regional formula-block mapping logic from the March
builder, but points it at the June-2026 vintage and preserves March outputs as a
separate historical layer. Exact regional subtotal reconciliation is mandatory.
"""
from __future__ import annotations

import csv
import io
import urllib.request

import openpyxl

import build_subtel_fixed_commune_2026 as base

URL = "https://www.subtel.gob.cl/wp-content/uploads/2026/08/1_SERIES_CONEXIONES_INTERNET_FIJA_JUN26_100826.xlsx"
SHEETS = {
    "total_fixed_connections_2026m06": "7.11.CO_FIJAS_COMUNA",
    "residential_fixed_connections_2026m06": "7.11.1.CO_FIJAS_RES_COMUNA",
}
PERIOD = "2026-06"
EXPECTED_NATIONAL_TOTAL = 4_900_369


def locate_period_column(ws):
    current_year = None
    matches = []
    for col in range(1, ws.max_column + 1):
        y = ws.cell(8, col).value
        if isinstance(y, (int, float)) and int(y) == y and 2000 <= int(y) <= 2100:
            current_year = int(y)
        month = base.clean(ws.cell(9, col).value).lower()
        if current_year == 2026 and month in {"jun", "junio"}:
            matches.append(col)
    if not matches:
        raise RuntimeError("Could not locate June 2026 column")
    return max(matches)


def main():
    request = urllib.request.Request(URL, headers={"User-Agent": "Mozilla/5.0 Chile-Digital-Inclusion/1.0"})
    with urllib.request.urlopen(request, timeout=120) as response:
        content = response.read()

    wb_values = openpyxl.load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    wb_formulas = openpyxl.load_workbook(io.BytesIO(content), read_only=False, data_only=False)

    with base.COMMUNES.open(encoding="utf-8") as f:
        catalogue = list(csv.DictReader(f))
    catalogue_by_region = {}
    for region in base.REGION_CODES:
        region_rows = [r for r in catalogue if int(r["region"]) == region]
        catalogue_by_region[region] = sorted(region_rows, key=lambda r: base.source_sort_key(r["comuna_nombre"]))

    all_alignment = []
    all_provenance = []
    all_issues = []
    metric_maps = {}
    cols = {}

    for metric, sheet_name in SHEETS.items():
        if sheet_name not in wb_values.sheetnames:
            raise RuntimeError(f"Expected sheet not found: {sheet_name}")
        ws_values = wb_values[sheet_name]
        ws_formulas = wb_formulas[sheet_name]
        target_col = locate_period_column(ws_values)
        cols[metric] = target_col
        mapped, alignment, provenance, issues = base.build_metric(
            ws_values, ws_formulas, target_col, catalogue_by_region, metric
        )
        metric_maps[metric] = mapped
        all_alignment.extend(alignment)
        all_provenance.extend(provenance)
        all_issues.extend(issues)

    if all_issues:
        raise RuntimeError(f"Regional formula-block reconciliation failed: {all_issues}")

    by_code = {int(r["comuna"]): r for r in catalogue}
    output_rows = []
    missing = []
    for code in sorted(by_code):
        r = by_code[code]
        total = metric_maps["total_fixed_connections_2026m06"].get(code)
        residential = metric_maps["residential_fixed_connections_2026m06"].get(code)
        share = round(residential / total * 100, 4) if total and residential is not None else None
        status = "reported" if total is not None and residential is not None else "source_blank"
        output_rows.append([
            r["region"], r["region_nombre"], r["provincia"], r["provincia_nombre"], r["comuna"], r["comuna_nombre"],
            total, residential, share, status, "regional_formula_block_order", PERIOD, URL,
        ])
        if status != "reported":
            missing.append([r["region"], r["region_nombre"], r["comuna"], r["comuna_nombre"], total, residential, status])

    with (base.OUT / "commune_fixed_connections_2026_06.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([
            "region", "region_nombre", "provincia", "provincia_nombre", "comuna", "comuna_nombre",
            "fixed_connections_total", "fixed_connections_residential", "residential_share_pct", "source_status",
            "source_mapping_method", "period", "source_url",
        ])
        w.writerows(output_rows)

    with (base.OUT / "source_not_reported_communes_2026_06.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([
            "region", "region_nombre", "comuna", "comuna_nombre", "fixed_connections_total",
            "fixed_connections_residential", "source_status",
        ])
        w.writerows(missing)

    alignment_fields = [
        "metric", "region", "block_start_row", "block_end_row", "block_rows", "regional_subtotal_row",
        "mapped_commune_slots", "catalogue_communes", "source_blank_slots", "regional_subtotal", "mapped_value_sum",
        "subtotal_delta", "source_labels_matching_mapped_names", "label_order_status", "count_status", "subtotal_status",
    ]
    base.write_dict_csv(base.OUT / "source_alignment_qa_2026_06.csv", alignment_fields, all_alignment)

    provenance_fields = [
        "metric", "region", "source_row", "source_commune_label", "mapped_commune", "mapped_commune_name", "value",
        "source_cell_status", "mapping_method", "block_start_row", "block_end_row", "regional_subtotal_row",
    ]
    base.write_dict_csv(base.OUT / "source_row_mapping_2026_06.csv", provenance_fields, all_provenance)

    with (base.OUT / "extraction_manifest_2026_06.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["metric", "source_sheet", "june_2026_column_number", "mapping_method", "source_url"])
        for metric, sheet_name in SHEETS.items():
            w.writerow([metric, sheet_name, cols[metric], "regional_formula_block_order_with_subtotal_reconciliation", URL])

    total_map = metric_maps["total_fixed_connections_2026m06"]
    residential_map = metric_maps["residential_fixed_connections_2026m06"]
    total_count = len(total_map)
    residential_count = len(residential_map)
    missing_codes = {int(r[2]) for r in missing}
    national_numeric_sum = sum(total_map.values())

    print(f"communes total fixed numeric: {total_count}/346")
    print(f"communes residential fixed numeric: {residential_count}/346")
    print(f"explicit source blanks: {len(missing)} {sorted(missing_codes)}")
    print(f"regional alignment checks: {len(all_alignment)} all pass")
    print(f"national fixed sum: {national_numeric_sum}")

    if total_count != 345 or residential_count != 345 or missing_codes != base.EXPECTED_SOURCE_BLANK_CODES:
        raise SystemExit("Unexpected formula-block reconstruction coverage")
    if national_numeric_sum != EXPECTED_NATIONAL_TOTAL:
        raise SystemExit(f"June commune total {national_numeric_sum} != national total {EXPECTED_NATIONAL_TOTAL}")


if __name__ == "__main__":
    main()
