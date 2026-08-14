from __future__ import annotations

import csv
import io
import re
from pathlib import Path

import openpyxl
import requests

SHEET_ID = '11EPq1k3dRWyrXfsWJWh405slQ-tyEjJj'
SOURCE_PAGE = 'https://www.innovacion.mineduc.cl/convocatoria-2025'
SOURCE_SHEET = f'https://docs.google.com/spreadsheets/d/{SHEET_ID}/edit'
OUT = Path('data/education_connectivity_2026/aulas_conectadas_2025_sheet_profile.csv')
HEADERS = Path('data/education_connectivity_2026/aulas_conectadas_2025_sheet_headers.csv')


def clean(v):
    return re.sub(r'\s+', ' ', '' if v is None else str(v)).strip()


def download():
    url=f'https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx'
    r=requests.get(url,timeout=120,allow_redirects=True); r.raise_for_status()
    if not r.content.startswith(b'PK'):
        raise RuntimeError(f'Unexpected workbook payload type={r.headers.get("content-type")} bytes={len(r.content)}')
    return r.content


def score_header(values):
    joined=' | '.join(clean(v).upper() for v in values if clean(v))
    terms=['RBD','ESTABLEC','NOMBRE','COMUNA','REGION','REGIÓN','ESTADO','SITUACION','SITUACIÓN','SELECCION','SELECCIÓN']
    return sum(t in joined for t in terms)


def main():
    OUT.parent.mkdir(parents=True,exist_ok=True)
    body=download(); wb=openpyxl.load_workbook(io.BytesIO(body),read_only=True,data_only=True)
    profile=[]; header_rows=[]
    for ws in wb.worksheets:
        preview=[]
        for i,row in enumerate(ws.iter_rows(min_row=1,max_row=min(ws.max_row,20),values_only=True),start=1):
            preview.append((i,list(row)))
        best=max(preview,key=lambda x: score_header(x[1])) if preview else (1,[])
        best_score=score_header(best[1])
        profile.append({
            'sheet':ws.title,'max_row':ws.max_row,'max_column':ws.max_column,
            'candidate_header_row':best[0],'header_score':best_score,
            'workbook_bytes':len(body),'source_page':SOURCE_PAGE,'source_sheet':SOURCE_SHEET,
        })
        for col_idx,value in enumerate(best[1],start=1):
            if clean(value):
                header_rows.append({'sheet':ws.title,'header_row':best[0],'column':col_idx,'header':clean(value),'header_score':best_score})

    with OUT.open('w',encoding='utf-8',newline='') as fh:
        fields=list(profile[0].keys()); w=csv.DictWriter(fh,fieldnames=fields); w.writeheader(); w.writerows(profile)
    with HEADERS.open('w',encoding='utf-8',newline='') as fh:
        fields=['sheet','header_row','column','header','header_score']; w=csv.DictWriter(fh,fieldnames=fields); w.writeheader(); w.writerows(header_rows)
    print('sheets',len(profile),'profiles',profile)

if __name__=='__main__':
    main()
